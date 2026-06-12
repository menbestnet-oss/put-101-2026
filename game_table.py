#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Путь 101% 2026 — Чтение игровой таблицы (Excel)

Источник истины по баллам игры: лист на каждый месяц (Апрель, Май, Июнь...).
Структура листа:
  - строки 1-2: шапка (даты, под каждой датой 3 колонки: утро / план дня / вечер)
  - строки 3-11: менеджеры с баллами по дням, колонка B = ИТОГ месяца
  - дальше блоки: недельные суммы / баллы за запуски / итоговый рейтинг

Результат кэшируется в game_cache.json — если Excel недоступен,
сайт собирается по последней удачной копии.
"""

import json
import os
from datetime import datetime, date

XLSX_PATH = r"C:\Users\Мансур\Downloads\Игра _Путь 101%_ 2026 год.xlsx"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
GAME_CACHE = os.path.join(SCRIPT_DIR, "game_cache.json")

MONTH_SHEETS = {
    'Январь': 1, 'Февраль': 2, 'Март': 3, 'Апрель': 4, 'Май': 5, 'Июнь': 6,
    'Июль': 7, 'Август': 8, 'Сентябрь': 9, 'Октябрь': 10, 'Ноябрь': 11, 'Декабрь': 12,
}


def _cell_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _parse_sheet(ws):
    """Один лист месяца → данные по менеджерам."""
    # Колонки дат: строка 1, под датой 3 колонки (утро/план дня/вечер)
    day_cols = {}  # col -> ISO date
    for col in range(3, ws.max_column + 1):
        d = _cell_date(ws.cell(row=1, column=col).value)
        if d:
            day_cols[col] = d.isoformat()

    managers = {}
    for row in range(3, 12):
        name = ws.cell(row=row, column=1).value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        total = ws.cell(row=row, column=2).value
        days = {}
        for col, iso in day_cols.items():
            def num(c):
                v = ws.cell(row=row, column=c).value
                return float(v) if isinstance(v, (int, float)) else 0.0
            m, p, e = num(col), num(col + 1), num(col + 2)
            if m or p or e:
                days[iso] = {'morning': m, 'plan': p, 'evening': e}
        managers[name] = {
            'total': float(total) if isinstance(total, (int, float)) else 0.0,
            'days': days,
        }

    # Бонус «План месяца» (101 балл) — блок с колонкой "План месяца"
    plan_bonus_col = None
    plan_header_row = None
    for row in range(12, min(ws.max_row, 60) + 1):
        for col in range(2, 10):
            v = ws.cell(row=row, column=col).value
            if isinstance(v, str) and 'план месяца' in v.lower():
                plan_bonus_col, plan_header_row = col, row
                break
        if plan_bonus_col:
            break
    if plan_bonus_col:
        for row in range(plan_header_row + 1, plan_header_row + 12):
            name = ws.cell(row=row, column=1).value
            if not name or not str(name).strip():
                continue
            name = str(name).strip()
            v = ws.cell(row=row, column=plan_bonus_col).value
            if name in managers and isinstance(v, (int, float)) and v > 0:
                managers[name]['plan_bonus'] = float(v)

    return managers


def read_game_table(log=print):
    """Читает Excel → {месяц 'YYYY-MM': {менеджер: {total, days, plan_bonus?}}}.

    При ошибке возвращает последний кэш. Результат сохраняется в game_cache.json.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
        data = {}
        for ws in wb.worksheets:
            month_num = MONTH_SHEETS.get(ws.title.strip())
            if not month_num:
                continue
            managers = _parse_sheet(ws)
            # Год определяем по датам листа
            year = 2026
            for m in managers.values():
                if m['days']:
                    year = int(next(iter(m['days'])).split('-')[0])
                    break
            data[f"{year}-{month_num:02d}"] = managers
        if data:
            tmp = GAME_CACHE + '.tmp'
            with open(tmp, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=1)
            os.replace(tmp, GAME_CACHE)
            log(f"Игровая таблица: {len(data)} мес. ({', '.join(sorted(data))})")
        return data
    except Exception as e:
        log(f"Excel недоступен ({e}) — использую кэш игровой таблицы")
        if os.path.exists(GAME_CACHE):
            with open(GAME_CACHE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}


if __name__ == '__main__':
    import sys
    if sys.platform == 'win32':
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    data = read_game_table()
    for month, managers in sorted(data.items()):
        print(f"\n=== {month} ===")
        for name, m in sorted(managers.items(), key=lambda x: -x[1]['total']):
            bonus = ' +план месяца' if m.get('plan_bonus') else ''
            print(f"  {name}: {m['total']:.0f} баллов, дней с баллами: {len(m['days'])}{bonus}")
